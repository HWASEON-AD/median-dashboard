# -*- coding: utf-8 -*-
"""
엑셀 파일 파싱 + Supabase 업로드 + 오늘 날짜 컬럼 추가 스크립트
사용법: python parse_excel.py "엑셀파일경로"
"""
import sys
import json
import openpyxl
from datetime import date, datetime

XLSX_PATH = sys.argv[1] if len(sys.argv) > 1 else None
if not XLSX_PATH:
    print('사용법: python parse_excel.py 엑셀파일경로')
    sys.exit(1)

wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))
headers = rows[0]

# 날짜 컬럼 위치 파악
date_cols = []
for i, h in enumerate(headers):
    if hasattr(h, 'year'):  # datetime 객체
        date_cols.append((i, h.strftime('%Y-%m-%d')))
    elif isinstance(h, str) and len(h) == 10 and h[4] == '-' and h[7] == '-':
        date_cols.append((i, h))

# 키워드 데이터 파싱
keywords = []
exposures = []  # {keyword, product, brand, date, is_exposed}

for row in rows[1:]:
    brand = str(row[0] or '').strip() or '아모스'
    product = str(row[1] or '').strip()
    keyword = str(row[2] or '').strip()
    tab = str(row[3] or '').strip()
    blog_url = str(row[4] or '').strip()
    hwaseon_url = str(row[5] or '').strip()

    if not keyword:
        continue

    keywords.append({
        'brand': brand,
        'product': product,
        'keyword': keyword,
        'tab': tab,
        'blog_url': blog_url or None,
        'hwaseon_url': hwaseon_url or None,
    })

    for col_idx, date_str in date_cols:
        val = row[col_idx] if col_idx < len(row) else None
        if val is None:
            continue
        val_str = str(val).strip()
        is_exposed = '노출' in val_str or val_str == '1' or val_str.lower() == 'true'
        is_not_exposed = val_str == '0' or val_str == '미노출'
        if is_exposed or is_not_exposed:
            exposures.append({
                'keyword': keyword,
                'product': product,
                'brand': brand,
                'date': date_str,
                'is_exposed': is_exposed,
            })

print(json.dumps({'keywords': keywords, 'exposures': exposures}, ensure_ascii=False))
