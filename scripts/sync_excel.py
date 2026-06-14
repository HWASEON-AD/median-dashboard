# -*- coding: utf-8 -*-
"""
sync_excel.py - 엑셀 -> Supabase 전체 동기화 (기존 데이터 교체)
형식: 브랜드|제품|키워드|노출탭|발행URL|제품링크URL|날짜1|날짜2|...
날짜 셀: '노출' = 노출, 비어있음/None = 미노출

실행: python scripts/sync_excel.py "엑셀파일경로"
"""
import sys, re, time, urllib.parse, io
from datetime import date, datetime
import openpyxl, requests

XLSX_PATH = sys.argv[1] if len(sys.argv) > 1 else None
SUPABASE_URL = os.environ.get('SUPABASE_URL', 'https://kepzsboxjulzygehmzpf.supabase.co')
SUPABASE_KEY = os.environ.get('SUPABASE_SERVICE_KEY', '')
TODAY = date.today().isoformat()
SB = {'apikey': SUPABASE_KEY, 'Authorization': f'Bearer {SUPABASE_KEY}', 'Content-Type': 'application/json'}

def log(msg):
    safe = msg.encode('utf-8', errors='replace').decode('utf-8')
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {safe}", flush=True)

# ── 엑셀 파싱 ──────────────────────────────────────────────────
def parse_excel(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]

    date_cols = []
    for i, h in enumerate(headers):
        if hasattr(h, 'year'):
            date_cols.append((i, h.strftime('%Y-%m-%d')))
        elif isinstance(h, str) and re.match(r'\d{4}-\d{2}-\d{2}', str(h)):
            date_cols.append((i, str(h)[:10]))

    keywords, exposures = [], []
    for row in rows[1:]:
        if not any(row): continue
        brand = str(row[0] or '').strip() or '아모스'
        product = str(row[1] or '').strip()
        keyword = str(row[2] or '').strip()
        tab = str(row[3] or '').strip()
        blog_url = str(row[4] or '').strip() or None
        hwaseon_url = str(row[5] or '').strip() or None
        if not keyword: continue

        keywords.append({'brand': brand, 'product': product, 'keyword': keyword,
                         'tab_type': tab, 'blog_url': blog_url, 'hwaseon_url': hwaseon_url})

        for ci, ds in date_cols:
            if ds == TODAY: continue
            val = row[ci] if ci < len(row) else None
            if val is None: continue
            v = str(val).strip()
            if not v: continue
            # '노출' 포함 = 노출, 그 외 값(미노출 등) = 스킵 (비노출 날짜는 기록 안 함)
            if '노출' in v and '미노출' not in v:
                exposures.append({'keyword': keyword, 'product': product, 'brand': brand, 'date': ds})

    log(f"파싱 완료: 키워드 {len(keywords)}개 / 노출기록(노출일만) {len(exposures)}건")
    return keywords, exposures

# ── Supabase ──────────────────────────────────────────────────
def delete_all():
    log("기존 데이터 전체 삭제 중...")
    # amos_daily_exposure 먼저 (FK)
    r1 = requests.delete(f'{SUPABASE_URL}/rest/v1/amos_daily_exposure?post_id=not.is.null', headers=SB, timeout=15)
    r2 = requests.delete(f'{SUPABASE_URL}/rest/v1/amos_posts?id=not.is.null', headers=SB, timeout=15)
    log(f"  exposure 삭제: {r1.status_code} / posts 삭제: {r2.status_code}")

def insert_keywords(keywords):
    log(f"키워드 {len(keywords)}개 삽입 중...")
    data = [{**k, 'status': '미노출'} for k in keywords]
    r = requests.post(f'{SUPABASE_URL}/rest/v1/amos_posts',
                      headers={**SB, 'Prefer': 'return=representation'}, json=data, timeout=20)
    if r.ok:
        inserted = r.json()
        log(f"  -> {len(inserted)}개 삽입 완료")
        return {f"{p['keyword']}|||{p.get('product') or ''}|||{p.get('brand') or '아모스'}": p['id'] for p in inserted}
    else:
        log(f"  insert 오류: {r.status_code} {r.text[:300]}")
        return {}

def insert_exposures(exposures, pm):
    log(f"노출기록 {len(exposures)}건 업로드 중...")
    seen, records, skip = set(), [], 0
    for e in exposures:
        key = f"{e['keyword']}|||{e.get('product') or ''}|||{e.get('brand') or '아모스'}"
        pid = pm.get(key)
        if not pid: skip += 1; continue
        uniq = f"{pid}|||{e['date']}"
        if uniq in seen: continue
        seen.add(uniq)
        records.append({'post_id': pid, 'date': e['date']})
    if skip: log(f"  매칭 실패 {skip}건 스킵")
    ok_count = 0
    for i in range(0, len(records), 500):
        batch = records[i:i+500]
        r = requests.post(f'{SUPABASE_URL}/rest/v1/amos_daily_exposure',
                          headers={**SB, 'Prefer': 'resolution=ignore-duplicates,return=minimal'},
                          json=batch, timeout=20)
        if r.ok: ok_count += len(batch)
        else: log(f"  배치 오류: {r.status_code} {r.text[:200]}")
    log(f"  -> {ok_count}건 저장")

# ── 오늘 네이버 검색 (requests) ───────────────────────────────
def parse_url(url):
    if not url: return {}
    n = url.replace("m.blog.naver.com","blog.naver.com").replace("m.cafe.naver.com","cafe.naver.com")
    m = re.search(r"blog\.naver\.com/([^/?#]+)/(\d+)", n)
    if m: return {'id': m.group(1), 'post_no': m.group(2)}
    m = re.search(r"cafe\.naver\.com/([^/?#]+)/(\d+)", n)
    if m: return {'id': m.group(1), 'post_no': m.group(2)}
    return {}

def check_naver(keyword, blog_url):
    p = parse_url(blog_url)
    post_no = p.get('post_no')
    blog_id = p.get('id')
    if not post_no: return False
    try:
        r = requests.get(
            f"https://m.search.naver.com/search.naver?query={urllib.parse.quote(keyword)}",
            headers={
                'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 16_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.0 Mobile/15E148 Safari/604.1',
                'Accept-Language': 'ko-KR,ko;q=0.9',
                'Referer': 'https://m.naver.com/',
            }, timeout=12
        )
        html = r.text
        return post_no in html and (not blog_id or blog_id in html)
    except Exception as e:
        log(f"  오류: {str(e)[:50]}")
        return False

def check_today(keywords, pm):
    log(f"오늘({TODAY}) 검색 체크 ({len([k for k in keywords if k.get('blog_url')])}개)...")
    exposed_records = []
    for i, kw in enumerate(keywords):
        if not kw.get('blog_url'): continue
        keyword = kw['keyword']
        key = f"{keyword}|||{kw.get('product') or ''}|||{kw.get('brand') or '아모스'}"
        pid = pm.get(key)
        if not pid: continue
        found = check_naver(keyword, kw['blog_url'])
        result = 'O 노출' if found else 'X 미노출'
        log(f"  [{i+1}] {keyword[:22]} -> {result}")
        if found:
            exposed_records.append({'post_id': pid, 'date': TODAY})
        time.sleep(0.8)

    if exposed_records:
        r = requests.post(f'{SUPABASE_URL}/rest/v1/amos_daily_exposure',
                          headers={**SB, 'Prefer': 'resolution=merge-duplicates,return=minimal'},
                          json=exposed_records, timeout=20)
        log(f"오늘 노출 {len(exposed_records)}건 저장 -> {'OK' if r.ok else r.text[:100]}")
    else:
        log("오늘 노출 확인된 키워드 없음")

# ── 메인 ─────────────────────────────────────────────────────
def main():
    if not XLSX_PATH:
        print("사용법: python scripts/sync_excel.py 엑셀파일경로")
        sys.exit(1)

    log(f"엑셀: {XLSX_PATH}")
    keywords, exposures = parse_excel(XLSX_PATH)

    delete_all()
    pm = insert_keywords(keywords)
    if not pm:
        log("키워드 삽입 실패, 종료")
        sys.exit(1)

    if exposures:
        insert_exposures(exposures, pm)

    check_today(keywords, pm)
    log("=== 완료 ===")

if __name__ == '__main__':
    main()
